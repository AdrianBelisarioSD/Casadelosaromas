import bcrypt
import json
import csv
from datetime import date

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Sum, F, Q
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from django.db import IntegrityError, transaction
from django.core.mail import send_mail
from django.conf import settings
import hmac


from .models import (
    EmpleadosLamerica, SolicitudesAcceso, SolicitudSoporte, Producto, Cliente, Venta,
    DetalleVenta, Movimiento, Proveedor, OrdenCompra, DetalleOrdenCompra, 
    Bodega, EmpleadoBodega
)


def usuario_es_administrador(request):
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return False
    empleado = EmpleadosLamerica.objects.filter(id_empleado=empleado_id).first()
    if not empleado or not empleado.cargo_empleado:
        return False
    return empleado.cargo_empleado.strip().lower() in ['administrador', 'dueño', 'dueno']

def usuario_es_dueno(request):
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return False
    empleado = EmpleadosLamerica.objects.filter(id_empleado=empleado_id).first()
    if not empleado or not empleado.cargo_empleado:
        return False
    return empleado.cargo_empleado.strip().lower() in ['dueño', 'dueno']

def login_view(request):
    if request.method == 'POST':
        correo = request.POST.get('correo')
        contrasena = request.POST.get('contrasena')

        empleado = EmpleadosLamerica.objects.filter(correo_de_empleado=correo).first()
        if empleado and bcrypt.checkpw(contrasena.encode('utf-8'), empleado.contrasena.encode('utf-8')):
            request.session['empleado_id'] = empleado.id_empleado
            return redirect('inicio')
        else:
            messages.error(request, 'Correo o contraseña incorrectos')

    return render(request, 'gesym/login.html')


def recuperar_view(request):
    if request.method == 'POST':
        correo = request.POST.get('correo')
        messages.success(request, 'Si el correo existe, te enviaremos las instrucciones pronto')
        return redirect('recuperar')

    return render(request, 'gesym/recuperar.html')


def solicitud_view(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre_completo')
        documento = request.POST.get('documento_identidad')
        numero_empleado = request.POST.get('numero_empleado')
        correo = request.POST.get('correo_electronico')
        tipo = request.POST.get('cargo')
        mensaje_usuario = request.POST.get('mensaje')

        mensaje_completo = (
            f'Documento: {documento}\n'
            f'N° Empleado: {numero_empleado}\n\n'
            f'{mensaje_usuario}'
        )

        SolicitudSoporte.objects.create(
            nombre_completo=nombre,
            correo_electronico=correo,
            tipo=tipo,
            mensaje=mensaje_completo,
        )

        tipo_legible = dict(SolicitudSoporte.TIPO_CHOICES).get(tipo, tipo)
        send_mail(
            subject=f'Nuevo reporte de {tipo_legible} - MEKATARIO',
            message=(
                f'Tipo: {tipo_legible}\n'
                f'Nombre: {nombre}\n'
                f'Correo: {correo}\n\n'
                f'Mensaje:\n{mensaje_completo}'
            ),
            from_email=settings.EMAIL_HOST_USER,
            recipient_list=settings.CORREOS_ADMINISTRADORES,
            fail_silently=False,
        )

        messages.success(request, 'Tu solicitud fue enviada. El administrador te contactará pronto.')
        return redirect('login')

    return render(request, 'gesym/solicitud.html')


def registro_view(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre_completo')
        documento = request.POST.get('documento_identidad')
        numero_empleado = request.POST.get('numero_empleado')
        correo = request.POST.get('correo_electronico')
        cargo = request.POST.get('cargo')
        contraseña = request.POST.get('mensaje')

        SolicitudesAcceso.objects.create(
            nombre_completo=nombre,
            documento_identidad=documento,
            numero_empleado=numero_empleado,
            correo_electronico=correo,
            cargo=cargo,
            mensaje=contraseña,
        )

        link_solicitudes = request.build_absolute_uri('/solicitudes-pendientes/')
        send_mail(
            subject='Nueva solicitud de usuario de ingreso al sistema creado',
            message=(
                f'Se ha creado una nueva solicitud de usuario para ingresar al sistema MEKATARIO.\n\n'
                f'Nombre: {nombre}\n'
                f'Documento: {documento}\n'
                f'N° Empleado: {numero_empleado}\n'
                f'Correo: {correo}\n'
                f'Cargo: {cargo}\n\n'
                f'Para aprobar o rechazar esta solicitud, inicia sesión en el sistema y ve a:\n{link_solicitudes}'
            ),
            from_email=settings.EMAIL_HOST_USER,
            recipient_list=settings.CORREOS_ADMINISTRADORES,
            fail_silently=False,
        )

        messages.success(request, 'Tu solicitud fue enviada. El administrador aprobará tu cuenta pronto.')
        return redirect('login')

    return render(request, 'gesym/registro.html')


def solicitudes_pendientes_view(request):
    if not request.session.get('empleado_id'):
        return redirect('login')
    if not usuario_es_administrador(request):
        messages.error(request, 'No tienes permiso para acceder a esta sección.')
        return redirect('inicio')

    solicitudes = SolicitudesAcceso.objects.filter(
        Q(atendida__isnull=True) | Q(atendida=0)
    ).order_by('-fecha_solicitud')

    return render(request, 'gesym/solicitudes_pendientes.html', {'solicitudes': solicitudes})


def aprobar_solicitud_view(request, solicitud_id):
    if not request.session.get('empleado_id'):
        return redirect('login')
    if not usuario_es_administrador(request):
        messages.error(request, 'No tienes permiso para realizar esta acción.')
        return redirect('inicio')

    solicitud = get_object_or_404(SolicitudesAcceso, id_solicitud=solicitud_id)

    if request.method == 'POST':
        if EmpleadosLamerica.objects.filter(correo_de_empleado=solicitud.correo_electronico).exists():
            messages.error(request, f'Ya existe una cuenta con el correo {solicitud.correo_electronico}.')
            solicitud.atendida = 1
            solicitud.save()
            return redirect('solicitudes_pendientes')

        password_texto = solicitud.mensaje or 'mekatario123'
        hashed = bcrypt.hashpw(password_texto.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

        EmpleadosLamerica.objects.create(
            cargo_empleado=solicitud.cargo,
            nombre_empleado=solicitud.nombre_completo,
            apellido_empleado='',
            correo_de_empleado=solicitud.correo_electronico,
            direccion_empleado='',
            documento_de_identidad=int(solicitud.documento_identidad),
            contrasena=hashed,
        )

        solicitud.atendida = 1
        solicitud.save()
        messages.success(request, f'Cuenta creada correctamente para {solicitud.nombre_completo}.')

    return redirect('solicitudes_pendientes')


def rechazar_solicitud_view(request, solicitud_id):
    if not request.session.get('empleado_id'):
        return redirect('login')
    if not usuario_es_administrador(request):
        messages.error(request, 'No tienes permiso para realizar esta acción.')
        return redirect('inicio')

    solicitud = get_object_or_404(SolicitudesAcceso, id_solicitud=solicitud_id)

    if request.method == 'POST':
        solicitud.atendida = 2
        solicitud.save()
        messages.success(request, 'Solicitud rechazada correctamente.')

    return redirect('solicitudes_pendientes')


def soporte_view(request):
    empleado = None
    empleado_id = request.session.get('empleado_id')
    if empleado_id:
        empleado = EmpleadosLamerica.objects.filter(id_empleado=empleado_id).first()

    if request.method == 'POST':
        nombre = request.POST.get('nombre_completo')
        correo = request.POST.get('correo_electronico')
        tipo = request.POST.get('tipo')
        mensaje = request.POST.get('mensaje')

        SolicitudSoporte.objects.create(
            nombre_completo=nombre,
            correo_electronico=correo,
            tipo=tipo,
            mensaje=mensaje,
        )

        tipo_legible = dict(SolicitudSoporte.TIPO_CHOICES).get(tipo, tipo)
        send_mail(
            subject=f'Nuevo reporte de {tipo_legible} - MEKATARIO',
            message=(
                f'Tipo: {tipo_legible}\n'
                f'Nombre: {nombre}\n'
                f'Correo: {correo}\n\n'
                f'Mensaje:\n{mensaje}'
            ),
            from_email=settings.EMAIL_HOST_USER,
            recipient_list=settings.CORREOS_ADMINISTRADORES,
            fail_silently=False,
        )

        messages.success(request, 'Tu reporte fue enviado. El administrador lo revisará pronto.')
        return redirect('inicio') if empleado_id else redirect('login')

    return render(request, 'gesym/soporte.html', {'empleado': empleado})


def soporte_pendientes_view(request):
    if not request.session.get('empleado_id'):
        return redirect('login')
    if not usuario_es_administrador(request):
        messages.error(request, 'No tienes permiso para acceder a esta sección.')
        return redirect('inicio')

    reportes = SolicitudSoporte.objects.filter(
        Q(atendida__isnull=True) | Q(atendida=0)
    ).order_by('-fecha_solicitud')

    return render(request, 'gesym/soporte_pendientes.html', {'reportes': reportes})


def atender_soporte_view(request, solicitud_id):
    if not request.session.get('empleado_id'):
        return redirect('login')
    if not usuario_es_administrador(request):
        messages.error(request, 'No tienes permiso para realizar esta acción.')
        return redirect('inicio')

    reporte = get_object_or_404(SolicitudSoporte, id_solicitud=solicitud_id)

    if request.method == 'POST':
        reporte.atendida = 1
        reporte.save()
        messages.success(request, 'Reporte marcado como atendido.')

    return redirect('soporte_pendientes')


def descartar_soporte_view(request, solicitud_id):
    if not request.session.get('empleado_id'):
        return redirect('login')
    if not usuario_es_administrador(request):
        messages.error(request, 'No tienes permiso para realizar esta acción.')
        return redirect('inicio')

    reporte = get_object_or_404(SolicitudSoporte, id_solicitud=solicitud_id)

    if request.method == 'POST':
        reporte.atendida = 2
        reporte.save()
        messages.success(request, 'Reporte descartado.')

    return redirect('soporte_pendientes')


ROLES_PRIVILEGIADOS = ['administrador', 'dueño', 'dueno']


def usuarios_view(request):
    if not request.session.get('empleado_id'):
        return redirect('login')
    if not usuario_es_administrador(request):
        messages.error(request, 'No tienes permiso para acceder a esta sección.')
        return redirect('inicio')

    empleado_actual_id = request.session.get('empleado_id')
    es_dueno = usuario_es_dueno(request)

    usuarios = EmpleadosLamerica.objects.all().order_by('nombre_empleado')
    todas_bodegas = Bodega.objects.filter(activa=True).order_by('nombre')

    for u in usuarios:
        rol_actual = (u.cargo_empleado or '').strip().lower()
        u.rol_privilegiado = rol_actual in ROLES_PRIVILEGIADOS
        u.puede_editar_rol = (u.id_empleado != empleado_actual_id) and (not u.rol_privilegiado or es_dueno)
        u.es_rol_dueno = rol_actual in ['dueño', 'dueno']
        
        if u.rol_privilegiado:
            u.bodegas_ids = list(todas_bodegas.values_list('id', flat=True))
            u.bodegas_nombres = 'Todas las bodegas'
        else:
            bodegas_usuario = Bodega.objects.filter(empleados_asignados__empleado=u, activa=True).order_by('nombre')
            u.bodegas_ids = list(bodegas_usuario.values_list('id', flat=True))
            u.bodegas_nombres = ', '.join(bodegas_usuario.values_list('nombre', flat=True)) or '—'

    return render(request, 'gesym/usuarios.html', {
        'usuarios': usuarios,
        'empleado_actual_id': empleado_actual_id,
        'es_dueno': es_dueno,
        'todas_bodegas': todas_bodegas,
    })


def eliminar_usuario_view(request, usuario_id):
    if not request.session.get('empleado_id'):
        return redirect('login')
    if not usuario_es_administrador(request):
        messages.error(request, 'No tienes permiso para realizar esta acción.')
        return redirect('inicio')

    if request.session.get('empleado_id') == usuario_id:
        messages.error(request, 'No puedes eliminar tu propia cuenta mientras tienes la sesión iniciada.')
        return redirect('usuarios')

    usuario = get_object_or_404(EmpleadosLamerica, id_empleado=usuario_id)

    
    usuario = get_object_or_404(EmpleadosLamerica, id_empleado=usuario_id)

    rol_actual = (usuario.cargo_empleado or '').strip().lower()
    if rol_actual in ['dueño', 'dueno']:
        messages.error(request, 'El Dueño no puede ser eliminado por ningún usuario.')
        return redirect('usuarios')
        
    
    if request.method == 'POST':
        usuario.delete()
        messages.success(request, 'Usuario eliminado correctamente.')

    return redirect('usuarios')


def editar_usuario_view(request, usuario_id):
    if not request.session.get('empleado_id'):
        return redirect('login')
    if not usuario_es_administrador(request):
        messages.error(request, 'No tienes permiso para realizar esta acción.')
        return redirect('inicio')

    if request.session.get('empleado_id') == usuario_id:
        messages.error(request, 'No puedes editar tu propio usuario desde este panel.')
        return redirect('usuarios')

    usuario = get_object_or_404(EmpleadosLamerica, id_empleado=usuario_id)
    rol_actual = (usuario.cargo_empleado or '').strip().lower()

    if rol_actual in ROLES_PRIVILEGIADOS and not usuario_es_dueno(request):
        messages.error(request, 'Solo el dueño puede editar a un administrador o a otro dueño.')
        return redirect('usuarios')

    if request.method == 'POST':
        usuario.nombre_empleado = request.POST.get('nombre_empleado')
        usuario.apellido_empleado = request.POST.get('apellido_empleado')
        usuario.correo_de_empleado = request.POST.get('correo_de_empleado')
        usuario.direccion_empleado = request.POST.get('direccion_empleado')
        usuario.cargo_empleado = request.POST.get('cargo_empleado')

        documento = request.POST.get('documento_de_identidad')
        if documento:
            try:
                usuario.documento_de_identidad = int(documento)
            except ValueError:
                messages.error(request, 'El documento de identidad debe ser un número.')
                return redirect('usuarios')

        usuario.save()

        nuevo_rol = (usuario.cargo_empleado or '').strip().lower()
        if nuevo_rol not in ROLES_PRIVILEGIADOS:
            bodegas_ids = request.POST.getlist('bodegas_ids')
            EmpleadoBodega.objects.filter(empleado=usuario).delete()
            for bid in bodegas_ids:
                EmpleadoBodega.objects.create(empleado=usuario, bodega_id=bid)

        messages.success(request, 'Usuario actualizado correctamente.')

    return redirect('usuarios')


def inicio_view(request):
    if not request.session.get('empleado_id'):
        return redirect('login')

    bodega = obtener_bodega_actual(request)
    if not bodega:
        return redirect('seleccionar_bodega')

    productos = Producto.objects.filter(bodega=bodega)
    total_productos = productos.count()
    unidades_en_stock = productos.aggregate(total=Sum('cantidad'))['total'] or 0
    valor_total_inventario = sum(p.valor_total for p in productos)
    stock_bajo = productos.filter(cantidad__lt=F('stock_minimo')).count()

    movimientos_recientes = Movimiento.objects.filter(bodega=bodega).select_related('producto').order_by('-fecha')[:10]

    compras_recientes = OrdenCompra.objects.select_related('proveedor').order_by('-fecha')[:5]

    hoy = date.today()
    movimientos_mes = Movimiento.objects.filter(bodega=bodega, fecha__year=hoy.year, fecha__month=hoy.month)

    entradas_por_dia = {}
    salidas_por_dia = {}
    for m in movimientos_mes:
        dia = m.fecha.day
        if m.cantidad >= 0:
            entradas_por_dia[dia] = entradas_por_dia.get(dia, 0) + m.cantidad
        else:
            salidas_por_dia[dia] = salidas_por_dia.get(dia, 0) + abs(m.cantidad)

    dias = list(range(1, 31))
    grafico_entradas = [entradas_por_dia.get(d, 0) for d in dias]
    grafico_salidas = [salidas_por_dia.get(d, 0) for d in dias]

    contexto = {
        'total_productos': total_productos,
        'unidades_en_stock': unidades_en_stock,
        'valor_total_inventario': valor_total_inventario,
        'stock_bajo': stock_bajo,
        'productos_por_vencer': 0,
        'errores_inventario': 0,
        'movimientos_recientes': movimientos_recientes,
        'compras_recientes': compras_recientes,
        'dias_json': json.dumps(dias),
        'grafico_entradas': json.dumps(grafico_entradas),
        'grafico_salidas': json.dumps(grafico_salidas),
        'bodega_actual': bodega,
    }
    return render(request, 'gesym/inicio.html', contexto)


def inventario_view(request):
    if not request.session.get('empleado_id'):
        return redirect('login')

    bodega = obtener_bodega_actual(request)
    if not bodega:
        return redirect('seleccionar_bodega')

    productos_lista = Producto.objects.filter(bodega=bodega)

    query = request.GET.get('q', '').strip()
    if query:
        productos_lista = productos_lista.filter(
            Q(nombre__icontains=query) |
            Q(codigo__icontains=query) |
            Q(marca__icontains=query) |
            Q(categoria__icontains=query)
        )

    orden = request.GET.get('orden', 'nombre_asc')
    campos_orden = {
        'nombre_asc': 'nombre',
        'nombre_desc': '-nombre',
        'codigo_asc': 'codigo',
        'codigo_desc': '-codigo',
    }
    productos_lista = productos_lista.order_by(campos_orden.get(orden, 'nombre'))

    total_productos = productos_lista.count()
    unidades_en_stock = productos_lista.aggregate(total=Sum('cantidad'))['total'] or 0
    valor_total_inventario = sum(p.valor_total for p in productos_lista)
    stock_bajo = productos_lista.filter(cantidad__lt=F('stock_minimo')).count()

    paginator = Paginator(productos_lista, 13)
    numero_pagina = request.GET.get('page')
    productos = paginator.get_page(numero_pagina)

    contexto = {
        'productos': productos,
        'total_productos': total_productos,
        'unidades_en_stock': unidades_en_stock,
        'valor_total_inventario': valor_total_inventario,
        'stock_bajo': stock_bajo,
        'bodega_actual': bodega,
        'orden_actual': orden,
    }
    return render(request, 'gesym/inventario.html', contexto)


def agregar_producto_view(request):
    if not request.session.get('empleado_id'):
        return redirect('login')

    bodega = obtener_bodega_actual(request)
    if not bodega:
        return redirect('seleccionar_bodega')

    es_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'

    if request.method == 'POST':
        tipo_ingreso = request.POST.get('tipo_ingreso') or 'unidad'

        unidades_por_bandeja = None
        numero_bandejas = None
        if tipo_ingreso == 'bandeja':
            unidades_por_bandeja = request.POST.get('unidades_por_bandeja') or None
            numero_bandejas = request.POST.get('numero_bandejas') or None

        try:
            producto = Producto.objects.create(
                bodega=bodega,
                codigo=request.POST.get('codigo'),
                nombre=request.POST.get('nombre'),
                marca=request.POST.get('marca'),
                categoria=request.POST.get('categoria'),
                cantidad=request.POST.get('cantidad') or 0,
                precio=request.POST.get('precio') or 0,
                descripcion=request.POST.get('descripcion', ''),
                tipo_ingreso=tipo_ingreso,
                unidades_por_bandeja=unidades_por_bandeja,
                numero_bandejas=numero_bandejas,
                unidades_por_caja=request.POST.get('unidades_por_caja') or None,
                precio_caja=request.POST.get('precio_caja') or None,
            )
            registrar_movimiento(producto, 'Entrada', 'Ingreso de producto', producto.cantidad, request)

            if es_ajax:
                return JsonResponse({'ok': True, 'mensaje': 'Producto agregado correctamente'})
            messages.success(request, 'Producto agregado correctamente')

        except IntegrityError:
            if es_ajax:
                return JsonResponse(
                    {'ok': False, 'error': 'Ya existe un producto con ese código en esta bodega'},
                    status=400
                )
            messages.error(request, 'Ya existe un producto con ese código en esta bodega')

    return redirect('inventario')


def editar_producto_view(request, producto_id):
    if not request.session.get('empleado_id'):
        return redirect('login')

    bodega = obtener_bodega_actual(request)
    producto = get_object_or_404(Producto, id=producto_id, bodega=bodega)

    if request.method == 'POST':
        producto.codigo = request.POST.get('codigo')
        producto.nombre = request.POST.get('nombre')
        producto.marca = request.POST.get('marca')
        producto.categoria = request.POST.get('categoria')
        producto.cantidad = request.POST.get('cantidad') or 0
        producto.precio = request.POST.get('precio') or 0
        producto.descripcion = request.POST.get('descripcion', '')
        producto.tipo_ingreso = request.POST.get('tipo_ingreso') or 'unidad'
        producto.unidades_por_caja = request.POST.get('unidades_por_caja') or None
        producto.precio_caja = request.POST.get('precio_caja') or None
        try:
            producto.save()
            messages.success(request, 'Producto actualizado correctamente')
        except IntegrityError:
            messages.error(request, 'Ya existe un producto con ese código en esta bodega')

    return redirect('inventario')


def eliminar_producto_view(request, producto_id):
    if not request.session.get('empleado_id'):
        return redirect('login')
    
    bodega = obtener_bodega_actual(request)
    producto = get_object_or_404(Producto, id=producto_id, bodega=bodega)

    if request.method == 'POST':
        producto.delete()
        messages.success(request, 'Producto eliminado correctamente')

    return redirect('inventario')


def compras_view(request):
    if not request.session.get('empleado_id'):
        return redirect('login')
    
    bodega = obtener_bodega_actual(request)   
    if not bodega:                            
        return redirect('seleccionar_bodega') 

    productos_bodega = Producto.objects.filter(bodega=bodega).values_list('id', flat=True)
    ventas_lista = Venta.objects.select_related('cliente').prefetch_related('detalles').filter(
        detalles__producto_id__in=productos_bodega
    ).distinct().order_by('-fecha_venta', '-id')

    query = request.GET.get('q', '').strip()
    if query:
        ventas_lista = ventas_lista.filter(
            Q(cliente__nombre_completo__icontains=query) |
            Q(detalles__producto__nombre__icontains=query)
        ).distinct()

    total_ventas = sum(v.total_venta for v in ventas_lista)
    total_ventas_realizadas = ventas_lista.count()
    total_clientes = Cliente.objects.filter(bodega=bodega).count()

    paginator = Paginator(ventas_lista, 10)
    numero_pagina = request.GET.get('page')
    ventas = paginator.get_page(numero_pagina)
    clientes_json = list(Cliente.objects.filter(bodega=bodega).values('id', 'nombre_completo', 'correo', 'telefono', 'direccion'))
    productos_json = list(Producto.objects.filter(bodega=bodega).values(
        'id', 'nombre', 'codigo', 'precio', 'cantidad', 'tipo_ingreso', 'unidades_por_bandeja',
        'unidades_por_caja', 'precio_caja'
    ))

    contexto = {
    'ventas': ventas,
    'total_ventas': total_ventas,
    'total_ventas_realizadas': total_ventas_realizadas,
    'total_clientes': total_clientes,
    'productos_json': productos_json,
    'clientes_json': clientes_json,
    'bodega_actual': bodega,
    }   
    return render(request, 'gesym/compras.html', contexto)


def agregar_venta_view(request):
    if not request.session.get('empleado_id'):
        return redirect('login')
    
    bodega = obtener_bodega_actual(request)
    if not bodega:                            
        return redirect('seleccionar_bodega')

    if request.method == 'POST':
        productos = json.loads(request.POST.get('productos_json', '[]'))

        errores_stock = []
        for p in productos:
            prod = Producto.objects.filter(id=p.get('id'), bodega=bodega).first()
            if not prod:
                errores_stock.append(f"El producto '{p.get('nombre')}' ya no existe.")
            elif p.get('cantidad', 0) > prod.cantidad:
                errores_stock.append(
                    f"No hay suficiente stock de '{prod.nombre}': disponible {prod.cantidad}, solicitado {p.get('cantidad')}."
                )

        if errores_stock:
            for error in errores_stock:
                messages.error(request, error)
            return redirect('compras')

        cliente, _ = Cliente.objects.get_or_create(
            correo=request.POST.get('correo') or None,
            bodega=bodega,
            defaults={'nombre_completo': request.POST.get('nombre_completo')}
        )
        cliente.bodega = bodega
        cliente.nombre_completo = request.POST.get('nombre_completo')
        cliente.telefono = request.POST.get('telefono')
        cliente.direccion = request.POST.get('direccion')
        cliente.save()

        venta = Venta.objects.create(
            cliente=cliente,
            fecha_venta=request.POST.get('fecha_venta'),
            metodo_pago=request.POST.get('metodo_pago'),
            notas=request.POST.get('notas', ''),
        )

        for p in productos:
            DetalleVenta.objects.create(
                venta=venta,
                producto_id=p.get('id'),
                nombre_producto=p.get('nombre'),
                codigo_producto=p.get('codigo'),
                cantidad=p.get('cantidad'),
                precio_unitario=p.get('precio'),
            )

            prod = Producto.objects.filter(id=p.get('id')).first()
            if prod:
                prod.cantidad = max(0, prod.cantidad - p.get('cantidad'))
                prod.save()
                registrar_movimiento(prod, 'Salida', 'Venta', -p.get('cantidad'), request)

        messages.success(request, 'Venta registrada correctamente')

    return redirect('compras')


def editar_venta_view(request, venta_id):
    if not request.session.get('empleado_id'):
        return redirect('login')

    venta = get_object_or_404(Venta, id=venta_id)

    if request.method == 'POST':
        venta.cliente.nombre_completo = request.POST.get('nombre_completo')
        venta.cliente.correo = request.POST.get('correo')
        venta.cliente.telefono = request.POST.get('telefono')
        venta.cliente.direccion = request.POST.get('direccion')
        venta.cliente.save()

        venta.fecha_venta = request.POST.get('fecha_venta')
        venta.metodo_pago = request.POST.get('metodo_pago')
        venta.notas = request.POST.get('notas', '')
        venta.save()

        venta.detalles.all().delete()
        productos = json.loads(request.POST.get('productos_json', '[]'))
        for p in productos:
            DetalleVenta.objects.create(
                venta=venta,
                producto_id=p.get('id'),
                nombre_producto=p.get('nombre'),
                codigo_producto=p.get('codigo'),
                cantidad=p.get('cantidad'),
                precio_unitario=p.get('precio'),
            )

        messages.success(request, 'Venta actualizada correctamente')

    return redirect('compras')


def eliminar_venta_view(request, venta_id):
    if not request.session.get('empleado_id'):
        return redirect('login')

    venta = get_object_or_404(Venta, id=venta_id)

    if request.method == 'POST':
        for detalle in venta.detalles.all():
            if detalle.producto:
                detalle.producto.cantidad += detalle.cantidad
                detalle.producto.save()
                registrar_movimiento(
                    detalle.producto,
                    'Entrada',
                    'Venta eliminada',
                    detalle.cantidad,
                    request
                )

        cliente = venta.cliente
        venta.delete()

        # Si el cliente no tiene ninguna otra venta, se elimina también
        if cliente and not Venta.objects.filter(cliente=cliente).exists():
            cliente.delete()

        messages.success(request, 'Venta eliminada correctamente')

    return redirect('compras')


def registrar_movimiento(producto, tipo, motivo, cantidad, request):
    empleado_id = request.session.get('empleado_id')
    usuario = ''
    if empleado_id:
        emp = EmpleadosLamerica.objects.filter(id_empleado=empleado_id).first()
        if emp:
            usuario = f"{emp.nombre_empleado} {emp.apellido_empleado}"
    Movimiento.objects.create(
        producto=producto,
        bodega=producto.bodega, 
        nombre_producto=producto.nombre,
        codigo_producto=producto.codigo,
        tipo=tipo,
        motivo=motivo,
        cantidad=cantidad,
        usuario_nombre=usuario,
    )


def informes_view(request):
    if not request.session.get('empleado_id'):
        return redirect('login')

    bodega = obtener_bodega_actual(request)
    if not bodega:
        return redirect('seleccionar_bodega')

    tab = request.GET.get('tab', 'informe')
    contexto = {'tab': tab, 'bodega_actual': bodega}

    if tab == 'informe':
        productos = Producto.objects.filter(bodega=bodega)
        productos_bodega = productos.values_list('id', flat=True)
        ventas = Venta.objects.select_related('cliente').filter(
            detalles__producto_id__in=productos_bodega
        ).distinct()

        total_vendido = sum(v.total_venta for v in ventas)
        productos_vendidos = sum(d.cantidad for v in ventas for d in v.detalles.all())
        alertas_stock_bajo = productos.filter(cantidad__lt=F('stock_minimo')).count()
        faltas_stock = productos.filter(cantidad=0).count()

        ventas_tabla = []
        for v in ventas.order_by('-fecha_venta')[:20]:
            for d in v.detalles.all():
                ventas_tabla.append({
                    'fecha': v.fecha_venta,
                    'producto': d.nombre_producto,
                    'vendedor': v.cliente.nombre_completo,
                    'cantidad': d.cantidad,
                    'total_vendida': d.subtotal,
                })

        grafico_labels = []
        grafico_datos = []
        ventas_por_fecha = {}
        for v in ventas:
            fecha_str = v.fecha_venta.strftime('%d %b')
            ventas_por_fecha[fecha_str] = ventas_por_fecha.get(fecha_str, 0) + float(v.total_venta)
        for fecha, total in sorted(ventas_por_fecha.items()):
            grafico_labels.append(fecha)
            grafico_datos.append(total)

        contexto.update({
            'total_vendido': total_vendido,
            'productos_vendidos': productos_vendidos,
            'alertas_stock_bajo': alertas_stock_bajo,
            'faltas_stock': faltas_stock,
            'ventas_tabla': ventas_tabla,
            'grafico_labels': json.dumps(grafico_labels),
            'grafico_datos': json.dumps(grafico_datos),
        })

    elif tab == 'inventario':
        productos_lista = Producto.objects.filter(bodega=bodega).order_by('nombre')
        total_productos = productos_lista.count()
        unidades_en_stock = productos_lista.aggregate(total=Sum('cantidad'))['total'] or 0
        valor_total_inventario = sum(p.valor_total for p in productos_lista)

        paginator = Paginator(productos_lista, 10)
        numero_pagina = request.GET.get('page')
        productos_inventario = paginator.get_page(numero_pagina)

        contexto.update({
            'productos_inventario': productos_inventario,
            'total_productos_inf': total_productos,
            'unidades_en_stock_inf': unidades_en_stock,
            'valor_total_inventario_inf': valor_total_inventario,
        })

    elif tab == 'movimientos':
        movimientos_lista = Movimiento.objects.filter(bodega=bodega).order_by('-fecha')

        paginator = Paginator(movimientos_lista, 10)
        numero_pagina = request.GET.get('page')
        movimientos = paginator.get_page(numero_pagina)

        contexto.update({
            'movimientos': movimientos,
        })

    elif tab == 'compras':
        productos_bodega = Producto.objects.filter(bodega=bodega).values_list('id', flat=True)
        ventas_lista = Venta.objects.select_related('cliente').filter(
            detalles__producto_id__in=productos_bodega
        ).distinct().order_by('-fecha_venta', '-id')

        buscar = request.GET.get('buscar', '').strip()
        if buscar:
            ventas_lista = ventas_lista.filter(cliente__nombre_completo__icontains=buscar)

        paginator = Paginator(ventas_lista, 10)
        numero_pagina = request.GET.get('page')
        ventas_compras = paginator.get_page(numero_pagina)

        contexto.update({
            'ventas_compras': ventas_compras,
            'buscar': buscar,
        })

    return render(request, 'gesym/informes.html', contexto)


def editar_producto_informe_view(request, producto_id):
    if not request.session.get('empleado_id'):
        return redirect('login')

    producto = get_object_or_404(Producto, id=producto_id)

    if request.method == 'POST':
        producto.cantidad = request.POST.get('cantidad') or 0
        producto.precio = request.POST.get('precio') or 0
        producto.activo = request.POST.get('activo') == 'Activo'
        producto.save()
        messages.success(request, 'Producto actualizado correctamente')

    return redirect('/informes/?tab=inventario')


def exportar_compras_view(request):
    if not request.session.get('empleado_id'):
        return redirect('login')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="compras.csv"'
    writer = csv.writer(response)
    writer.writerow(['Fecha', 'Numero', 'Proveedor', 'Productos', 'Valor Total', 'Estado'])

    for orden in OrdenCompra.objects.select_related('proveedor').all().order_by('-fecha'):
        writer.writerow([orden.fecha, orden.numero, orden.proveedor.nombre, orden.productos_texto, orden.total, orden.estado])

    return response


def exportar_compras_excel_view(request):
    if not request.session.get('empleado_id'):
        return redirect('login')

    ventas = Venta.objects.select_related('cliente').all().order_by('-fecha_venta')

    wb = Workbook()
    ws = wb.active
    ws.title = "Compras"

    headers = ['Fecha', 'Cliente', 'Productos', 'Metodo de pago', 'Total']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1451A9', end_color='1451A9', fill_type='solid')

    for venta in ventas:
        ws.append([
            venta.fecha_venta.strftime('%d/%m/%Y'),
            venta.cliente.nombre_completo,
            venta.total_productos,
            venta.get_metodo_pago_display(),
            float(venta.total_venta),
        ])

    for columna in ws.columns:
        largo_max = max((len(str(c.value)) if c.value else 0) for c in columna)
        ws.column_dimensions[columna[0].column_letter].width = largo_max + 4

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="compras.xlsx"'
    wb.save(response)
    return response


def acerca_view(request):
    if not request.session.get('empleado_id'):
        return redirect('login')
    return render(request, 'gesym/acerca.html')


def logout_view(request):
    request.session.flush()
    return redirect('login')


def obtener_bodega_actual(request):
    bodega_id = request.session.get('bodega_id')
    if not bodega_id:
        return None
    return Bodega.objects.filter(id=bodega_id, activa=True).first()


def seleccionar_bodega_view(request):
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return redirect('login')

    if usuario_es_administrador(request):
        bodegas = Bodega.objects.filter(activa=True).order_by('nombre')
    else:
        bodegas = Bodega.objects.filter(empleados_asignados__empleado_id=empleado_id, activa=True).order_by('nombre')

    if request.method == 'POST':
        bodega_id = request.POST.get('bodega_id')
        if bodegas.filter(id=bodega_id).exists():
            request.session['bodega_id'] = int(bodega_id)
            return redirect('inicio')
        messages.error(request, 'No tienes acceso a esa bodega.')

    if bodegas.count() == 1:
        request.session['bodega_id'] = bodegas.first().id
        return redirect('inicio')

    if bodegas.count() == 0:
        messages.error(request, 'No tienes ninguna bodega asignada. Contacta a un administrador.')

    return render(request, 'gesym/seleccionar_bodega.html', {'bodegas': bodegas})


def bodegas_view(request):
    if not request.session.get('empleado_id'):
        return redirect('login')
    bodegas = Bodega.objects.all().order_by('nombre')
    return render(request, 'gesym/bodegas.html', {
        'bodegas': bodegas,
        'es_dueno': usuario_es_dueno(request),
    })

def agregar_bodega_view(request):
    if not request.session.get('empleado_id'):
        return redirect('login')

    if request.method == 'POST':
        Bodega.objects.create(
            nombre=request.POST.get('nombre'),
            ubicacion=request.POST.get('ubicacion'),
            encargado=request.POST.get('encargado'),
        )
        messages.success(request, 'Bodega creada correctamente')

    return redirect('bodegas')


def editar_bodega_view(request, bodega_id):
    if not request.session.get('empleado_id'):
        return redirect('login')

    bodega = get_object_or_404(Bodega, id=bodega_id)

    if request.method == 'POST':
        bodega.nombre = request.POST.get('nombre')
        bodega.ubicacion = request.POST.get('ubicacion')
        bodega.encargado = request.POST.get('encargado')
        bodega.activa = request.POST.get('activa') == 'Activa'
        bodega.save()
        messages.success(request, 'Bodega actualizada correctamente')

    return redirect('bodegas')


def eliminar_bodega_view(request, bodega_id):
    if not request.session.get('empleado_id'):
        return redirect('login')

    bodega = get_object_or_404(Bodega, id=bodega_id)

    if request.method == 'POST':
        if bodega.productos.exists():
            messages.error(request, 'No puedes eliminar una bodega que tiene productos asignados')
        else:
            bodega.delete()
            messages.success(request, 'Bodega eliminada correctamente')

    return redirect('bodegas')

def clientes_view(request):
    if not request.session.get('empleado_id'):
        return redirect('login')

    bodega = obtener_bodega_actual(request)
    if not bodega:
        return redirect('seleccionar_bodega')

    clientes_lista = Cliente.objects.filter(bodega=bodega).prefetch_related('ventas').order_by('nombre_completo')

    query = request.GET.get('q', '').strip()
    if query:
        clientes_lista = clientes_lista.filter(
            Q(nombre_completo__icontains=query) |
            Q(correo__icontains=query) |
            Q(telefono__icontains=query)
        )

    total_clientes = clientes_lista.count()
    total_compras = sum(c.ventas.count() for c in clientes_lista)
    total_gastado = sum(v.total_venta for c in clientes_lista for v in c.ventas.all())

    paginator = Paginator(clientes_lista, 10)
    numero_pagina = request.GET.get('page')
    clientes = paginator.get_page(numero_pagina)

    contexto = {
        'clientes': clientes,
        'total_clientes': total_clientes,
        'total_compras': total_compras,
        'total_gastado': total_gastado,
        'bodega_actual': bodega,
    }
    return render(request, 'gesym/clientes.html', contexto)

def agregar_cliente_view(request):
    if not request.session.get('empleado_id'):
        return redirect('login')

    bodega = obtener_bodega_actual(request)
    if not bodega:
        return redirect('seleccionar_bodega')

    if request.method == 'POST':
        Cliente.objects.create(
            bodega=bodega,
            nombre_completo=request.POST.get('nombre_completo'),
            correo=request.POST.get('correo', ''),
            telefono=request.POST.get('telefono', ''),
            direccion=request.POST.get('direccion', ''),
        )
        messages.success(request, 'Cliente registrado correctamente')

    return redirect('clientes')

def editar_cliente_view(request, cliente_id):
    if not request.session.get('empleado_id'):
        return redirect('login')

    bodega = obtener_bodega_actual(request)
    cliente = get_object_or_404(Cliente, id=cliente_id, bodega=bodega)

    if request.method == 'POST':
        cliente.nombre_completo = request.POST.get('nombre_completo')
        cliente.correo = request.POST.get('correo', '')
        cliente.telefono = request.POST.get('telefono', '')
        cliente.direccion = request.POST.get('direccion', '')
        cliente.save()
        messages.success(request, 'Cliente actualizado correctamente')

    return redirect('clientes')

def eliminar_cliente_view(request, cliente_id):
    if not request.session.get('empleado_id'):
        return redirect('login')

    bodega = obtener_bodega_actual(request)
    cliente = get_object_or_404(Cliente, id=cliente_id, bodega=bodega)

    if request.method == 'POST':
        if cliente.ventas.exists():
            messages.error(request, 'No puedes eliminar un cliente que tiene ventas registradas.')
        else:
            cliente.delete()
            messages.success(request, 'Cliente eliminado correctamente')

    return redirect('clientes')


def reiniciar_bodega_view(request, bodega_id):
    if not request.session.get('empleado_id'):
        return redirect('login')

    if not usuario_es_dueno(request):
        messages.error(request, 'Solo el dueño puede reiniciar el inventario de una bodega.')
        return redirect('bodegas')

    bodega = get_object_or_404(Bodega, id=bodega_id)

    if request.method == 'POST':
        contrasena_confirmacion = request.POST.get('contrasena_confirmacion', '')

        if not hmac.compare_digest(contrasena_confirmacion, settings.RESET_BODEGA_PASSWORD):
            messages.error(request, 'Contraseña de confirmación incorrecta. No se reinició el inventario.')
            return redirect('bodegas')

        with transaction.atomic():
            productos_ids = list(Producto.objects.filter(bodega=bodega).values_list('id', flat=True))
            Venta.objects.filter(detalles__producto_id__in=productos_ids).distinct().delete()
            Cliente.objects.filter(bodega=bodega).delete()
            Movimiento.objects.filter(bodega=bodega).delete()
            Producto.objects.filter(bodega=bodega).delete()

        messages.success(request, f'El inventario de "{bodega.nombre}" se reinició correctamente.')

    return redirect('bodegas')